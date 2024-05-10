import openpyxl

#File --> Workbook --> Sheet --> Rows --> Cell

file="C:\\Users\\dhananjay.pokale\\Desktop\MyPersonal\\Python_DurgaSoft\\test_file.xlsx"
workbook=openpyxl.load_workbook((file))
sheet=workbook["Sheet1"]

num_rows=sheet.max_row
num_columns=sheet.max_column

#Example 1:
for r in range(1,num_rows+1):
    for c in range(1,num_columns+1):
        print(sheet.cell(r,c).value,end="   ")
    print()

#Example 2:
BookName="Java"

for r in range(1,num_rows+1):
    for c in range(1,num_columns+1):
        if sheet.cell(r,c).value==BookName:
            print("Price of the book is:",sheet.cell(r,3).value,end="   ")


#Example 3:
for r in range(1,num_rows+1):
    cell_value=sheet.cell(r,3).value
    if cell_value is not None:
        try:
            cell_value = int(cell_value)
            if int(cell_value) >= 200:
                print("Book Name is", sheet.cell(r, 1).value,"Whose Value is above or equal to 200", end="    ")
            print()
        except ValueError:
            pass


