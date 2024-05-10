import openpyxl

#Example 2:Write diffrent data
#Approach 1:
file="C:\\Users\\dhananjay.pokale\\Desktop\MyPersonal\\Python_DurgaSoft\\test_file.xlsx"
workbook=openpyxl.load_workbook((file))
sheet=workbook["Data2"]

sheet.cell(1,1).value='BookName'
sheet.cell(1,2).value='Price'
sheet.cell(1,3).value='Location'

sheet.cell(2,1).value='JAVA'
sheet.cell(2,2).value=200
sheet.cell(2,3).value='Pune'

sheet.cell(3,1).value='Python'
sheet.cell(3,2).value=300
sheet.cell(3,3).value='Mumbai'

workbook.save(file)